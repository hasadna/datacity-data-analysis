import re

from openpyxl import load_workbook
from xlrd import open_workbook

from config import CONFIGURATION, Config

MAGICS = [
    'סמל הרשות',
    '^מחוז',
    'מעמד מוניציפאלי',
    'מיסים ומענקים',
    'שם הרשות',
    'גירעון מצטבר בסוף שנה',
    'סמל רשות מקומית',
]
SPACES = re.compile('\s+')
LETTERS = re.compile('[א-ת]')
MAGICS = [
    re.compile(x) for x in MAGICS
]
MIN_SIZE = 30
HEADER_SIZE = 5


def clean(v):
    if isinstance(v, str):
        return SPACES.sub(' ', v.strip())
    return v


def get_safe(data, r, c):
    try:
        val = data[r][c]
        if val is not None:
            val = str(val).strip()
        return val
    except IndexError:
        return None


def fix_value(v, bad_floats):
    if v in ('-', '..', '', '.', None):
        return None
    if LETTERS.findall(v):
        return v
    v = v.replace(',', '')
    v = v.replace('*', '')
    v = v.strip()
    if v.startswith('(') and v.endswith(')'):
        v = v[1:-1]
    try:
        float(v)
    except Exception as e:
        bad_floats.add(v)
    return v


def process_sheet(year, name, data, filename): #noqa
    config = CONFIGURATION.get(year, dict()).get(name.strip()) or Config()
    # print(config)
    if config.skip:
        return

    magic_pos = []
    candidates = set()
    for row in range(HEADER_SIZE):
        # print('looking for', magic, 'in', row, 'row', data[row][:5])
        for col, val in enumerate(data[row]):
            if isinstance(val, str):
                candidates.add(((row, col), val))
    for row, rowdata in enumerate(data):
        for col, val in enumerate(rowdata[:HEADER_SIZE]):
            if isinstance(val, str):
                candidates.add(((row, col), val))

    for magic in MAGICS:
        for pos, val in candidates:
            if magic.search(val):
                magic_pos.append([*pos, magic])
                break

    assert len(magic_pos) >= 2, f'Failed to find enough magic positions {year}, {name}, {magic_pos}'
    # print(magic_pos, [data[r][c] for r, c, _ in magic_pos])

    if magic_pos[0][1] == magic_pos[1][1]:
        min_row = magic_pos[0][1]
        min_col = min(p[0] for p in magic_pos)
        get = lambda r, c: get_safe(data, c, r)
    elif magic_pos[0][0] == magic_pos[1][0]:
        min_row = magic_pos[0][0]
        min_col = min(p[1] for p in magic_pos)
        get = lambda r, c: get_safe(data, r, c)
    else:
        assert False, f'invalid magic positions {year}, {name}, {magic_pos}'

    headers_size = config.header_rows
    headers = []
    column = []
    headers_front = [None] * headers_size
    min_row = min_row - config.extend_headers_top

    # print('MMM', min_row, min_col, [get(min_row, min_col + i) for i in range(10)])

    c = 0
    name_idx = None
    while True:
        column = [
            get(min_row + r, min_col + c)
            for r in range(headers_size)
        ]
        if not any(column):
            break
        for i, v in enumerate(column):
            if v and not 'שנת עדכון' in v:
                headers_front[i] = v
                if len(LETTERS.findall(v))>0:
                    for ii in range(i + 1, headers_size):
                        headers_front[ii] = None
        # print(year, name, c, headers_front, column)
        # print(header)
        header = [v for v in headers_front if v]

        # print('HHH', headers_front)

        # Clear fully numeric headers from the front, they are not to be pulled through
        for i, v in enumerate(headers_front):
            if v and len(LETTERS.findall(v))==0:
                # print('CLEARING FRONT', c, repr(headers_front[i]))
                headers_front[i] = None

        if any(f in h for h in header for f in ('שם הרשות', 'סמל הרשות')):
            headers_front = [None] * headers_size            
        if any('שם הרשות' in h for h in header):
            assert name_idx is None, f'name_idx already set {name_idx}, {year}, {name}, {header}, {headers}'
            name_idx = len(headers)
        try:
            float(headers[-1])
            assert False, f'Bad Header extracted: {year}, {name}, {header}'
        except:
            pass
        header = [x.strip() for x in header if x]
        header = [x for x in header if x and len(x) > 1]
        header = '/'.join(header)
        headers.append(header)
        assert header != 'שכר ורווחה/2015/גברים', repr((header, headers_front))

        c += 1
    assert name_idx is not None, f'Failed to find name index {year}, {name}, {headers[:10]}'
    r = min_row + headers_size
    bad_floats = set()
    while True:
        row = [
            get(r, min_col + c)
            for c in range(len(headers))
        ]
        if not any(row):
            break
        muni_name = row[name_idx].strip().replace('*', '')
        if len([x for x in row if x]) < MIN_SIZE/2:
            break
        for h, v in zip(headers, row):
            if h != headers[name_idx]:
                v = fix_value(v, bad_floats)
                item = dict(
                    year=year,
                    # sheet=name,
                    name=muni_name,
                    header=h,
                    value=v,
                    filename=filename,
                )
                yield item
        r += 1
    if len(bad_floats) > 0:
        print('BAD FLOATS', bad_floats)


def preprocess_file(year, filename):
    if filename.endswith('.xlsx'):
        wb = load_workbook(filename)
        for name in wb.sheetnames:
            sheet = wb[name]
            # print(name, sheet.max_column, sheet.max_row)
            if sheet.max_column < MIN_SIZE or sheet.max_row < MIN_SIZE:
                continue
            data = [
                [clean(cell.value) for cell in row]
                for row in sheet.rows
            ]
            yield from process_sheet(year, sheet.title, data, filename)
    else:
        wb = open_workbook(filename)
        for idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(idx)
            if sheet.nrows < MIN_SIZE or sheet.ncols < MIN_SIZE:
                continue
            data = [
                [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            yield from process_sheet(year, sheet.name, data, filename)


def preprocess_files(filenames):
    for year, filename in filenames.items():
        print(year, filename)
        # if year < 2016: continue
        yield from preprocess_file(year, filename)
